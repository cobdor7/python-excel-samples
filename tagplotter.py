import os


from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'
i = 1
ws = wb.active

rootDir = "//USFL-S-fP01/productiontesters$/SWT/Products/G3/Production"
csvList = list()
pngList = list()
rcpList = list()


for dirName, subdirList, fileList in os.walk(rootDir):
    #print('Found directory: %s' % dirName)
    for fname in fileList:
        print('\t%s' % fname)
        if fname.endswith(".sup"):
            csvList.append(fname)
        if fname.endswith(".png"):
            pngList.append(fname)
        if fname.endswith(".rcp"):
            rcpList.append(fname)
            print(fname)
for tags in csvList:
    x = tags.find("_")
    if(x > 0):
        ws["A" + str(i)] = tags[:(x)]
        ws["B" + str(i)] = tags
        for png in pngList:
            if png[:-4] == tags[:x]:
                ws['C' + str(i)] = png
                print("FoundPNG!!")
                png.splitlines
        for rcp in rcpList:
            if rcp[:-4] == tags[:x]:
                ws['D' + str(i)] = rcp
                print(i)
        i = i + 1
print("Done!")
wb.save("newboy2.xlsx")
