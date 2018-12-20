import os
import csv



from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book5.xlsx'
i = 1
ws = wb.active


idComp = 'Product.Spec' #19
s1Comp = "Marker1.State" # 30
s2Comp = "Marker2.State" # 30
rootDir = "C:/Users/Ext.Ajames/Documents/Production"
csvList = list()
idlist = list()
s1list = list()
s2list = list()
print('5')
for dirName, subdirList, fileList in os.walk(rootDir):
    #print('Found directory: %s' % dirName)
    for fname in fileList:
        #print('\t%s' % fname)
        if fname.endswith(".csv"):
            csvList.append(fname)
            print(fname)
    
for tags in csvList:
    s1count = 0;
    s2count = 0;
    inputfile = csv.reader(open(rootDir + "/" + tags,'r'))
    for row in inputfile:
        fi = str(row)
        if(idComp in fi):
            idlist.append(fi[(fi.rfind(idComp)+len(idComp) + 2): -2])
            print(fi)
        if(s1Comp in fi):
            if not(s1count):
                s1list.append(fi[(fi.rfind(s1Comp)+len(s1Comp)+2): -2])
                s1count = 1
            print(fi)
        if(s2Comp in fi):    
            if not(s2count):
                s2list.append(fi[(fi.rfind(s2Comp)+len(s2Comp)+2): -2])
                s2count = 1
            print(fi)
      
i = 1
for tags in idlist:
    ws["A" + str(i)] = tags
    i = i + 1
i = 1
for png in s1list:
    ws['B' + str(i)] = png
    i = i + 1
i = 1
for rcp in s2list:
    ws['C' + str(i)] = rcp  
    i = i + 1
i=1
for rcp in csvList:
    ws['D' + str(i)] = rcp  
    i = i + 1
                
print("Done!")
wb.save("newboy3.xlsx")
